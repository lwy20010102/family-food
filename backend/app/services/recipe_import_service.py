from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models.recipe import Recipe
from app.models.recipe_import_backup import RecipeImportBackup
from app.schemas.recipe import (
    RecipeCreateRequest,
    RecipeIngredientBase,
    RecipeStepBase,
)
from app.schemas.recipe_import import (
    RecipeImportIssue,
    RecipeImportPreview,
    RecipeImportPreviewItem,
    RecipeImportResult,
    RecipeImportResultItem,
    RecipeImportUndoResult,
)
from app.services.recipe_service import (
    _apply_recipe_fields,
    _replace_ingredients,
    _replace_steps,
)


MAX_IMPORT_FILE_BYTES = 100 * 1024 * 1024
MAX_PREVIEW_RECIPES = 200
MAX_ISSUES = 200

RECIPE_CATEGORIES = {"肉类", "海鲜", "蔬菜", "主食", "汤", "早餐", "甜品", "其他"}
RECIPE_DIFFICULTIES = {"简单", "中等", "困难"}
RECIPE_INGREDIENT_TYPES = {"ingredient", "seasoning"}
RECIPE_SOURCE_TYPES = {"manual", "ai_text", "ai_video"}
RECIPE_STATUSES = {"草稿", "可导入"}

RECIPE_REQUIRED_HEADERS = {
    "recipe_key",
    "title",
    "category",
    "default_servings",
    "difficulty",
    "source_type",
    "status",
}
INGREDIENT_REQUIRED_HEADERS = {"recipe_key", "name", "amount", "unit", "type", "sort_order"}
STEP_REQUIRED_HEADERS = {"recipe_key", "step_number", "description"}


@dataclass(frozen=True)
class RecipeImportRecord:
    recipe_key: str
    data: RecipeCreateRequest


def preview_recipe_workbook(filename: str, file_bytes: bytes) -> RecipeImportPreview:
    """Parse and validate a recipe workbook without touching the database."""
    try:
        workbook = load_workbook(
            filename=BytesIO(file_bytes),
            read_only=True,
            data_only=False,
        )
    except Exception as exc:
        raise ValueError("无法读取 Excel 文件，请确认上传的是有效的 .xlsx 文件") from exc

    errors: list[RecipeImportIssue] = []
    warnings: list[RecipeImportIssue] = []
    sheets = workbook.sheetnames
    for required_sheet in ("菜谱", "食材", "步骤"):
        if required_sheet not in sheets:
            _add_issue(
                errors,
                RecipeImportIssue(
                    severity="error",
                    sheet=required_sheet,
                    message=f"缺少必需工作表：{required_sheet}",
                ),
            )

    if errors:
        return RecipeImportPreview(
            filename=filename,
            file_size_bytes=len(file_bytes),
            sheets=sheets,
            errors=errors,
            can_import=False,
        )

    recipe_rows, recipe_headers = _read_rows(workbook["菜谱"])
    ingredient_rows, ingredient_headers = _read_rows(workbook["食材"])
    step_rows, step_headers = _read_rows(workbook["步骤"])

    _check_headers("菜谱", recipe_headers, RECIPE_REQUIRED_HEADERS, errors)
    _check_headers("食材", ingredient_headers, INGREDIENT_REQUIRED_HEADERS, errors)
    _check_headers("步骤", step_headers, STEP_REQUIRED_HEADERS, errors)
    if errors:
        return RecipeImportPreview(
            filename=filename,
            file_size_bytes=len(file_bytes),
            sheets=sheets,
            errors=errors,
            can_import=False,
        )

    recipes: list[dict[str, Any]] = []
    recipes_by_key: dict[str, dict[str, Any]] = {}
    recipe_error_keys: set[str] = set()
    blank_tips = 0

    for row_number, row in recipe_rows:
        key = _text(row.get("recipe_key"))
        if not key:
            _add_issue(
                errors,
                RecipeImportIssue(
                    severity="error",
                    sheet="菜谱",
                    row=row_number,
                    field="recipe_key",
                    message="recipe_key 不能为空",
                ),
            )
            continue

        if key in recipes_by_key:
            recipe_error_keys.add(key)
            _add_issue(
                errors,
                RecipeImportIssue(
                    severity="error",
                    sheet="菜谱",
                    row=row_number,
                    field="recipe_key",
                    message=f"recipe_key 重复：{key}",
                ),
            )
            continue

        recipe = {
            "recipe_key": key,
            "title": _text(row.get("title")) or "",
            "description": _text(row.get("description")) or "",
            "category": _text(row.get("category")) or "",
            "image_url": _text(row.get("image_url")),
            "default_servings": _integer(row.get("default_servings")),
            "cooking_time": _integer(row.get("cooking_time")),
            "cooking_time_raw": _text(row.get("cooking_time")),
            "difficulty": _text(row.get("difficulty")) or "",
            "source_type": _text(row.get("source_type")) or "",
            "source_url": _text(row.get("source_url")),
            "status": _text(row.get("status")) or "",
            "data_note": _text(row.get("data_note")),
            "ingredient_count": 0,
            "step_count": 0,
        }
        recipes.append(recipe)
        recipes_by_key[key] = recipe
        if not _text(row.get("tips")):
            blank_tips += 1

        _validate_recipe_row(recipe, row_number, errors, warnings, recipe_error_keys)

    ingredient_count_by_key: dict[str, int] = {}
    ingredient_unspecified_amount = 0
    ingredient_blank_unit = 0
    for row_number, row in ingredient_rows:
        key = _text(row.get("recipe_key"))
        if not key:
            _add_issue(
                errors,
                RecipeImportIssue(
                    severity="error",
                    sheet="食材",
                    row=row_number,
                    field="recipe_key",
                    message="recipe_key 不能为空",
                ),
            )
            continue
        if key not in recipes_by_key:
            _add_issue(
                errors,
                RecipeImportIssue(
                    severity="error",
                    sheet="食材",
                    row=row_number,
                    field="recipe_key",
                    message=f"找不到对应的菜谱：{key}",
                ),
            )
            continue

        ingredient_count_by_key[key] = ingredient_count_by_key.get(key, 0) + 1
        name = _text(row.get("name"))
        if not name:
            recipe_error_keys.add(key)
            _add_issue(
                errors,
                RecipeImportIssue(
                    severity="error",
                    sheet="食材",
                    row=row_number,
                    field="name",
                    message="食材名称不能为空",
                ),
            )

        amount = _text(row.get("amount"))
        unit = _text(row.get("unit"))
        if not amount or "未标注" in amount:
            ingredient_unspecified_amount += 1
        if not unit:
            ingredient_blank_unit += 1

        ingredient_type = _text(row.get("type"))
        if ingredient_type not in RECIPE_INGREDIENT_TYPES:
            recipe_error_keys.add(key)
            _add_issue(
                errors,
                RecipeImportIssue(
                    severity="error",
                    sheet="食材",
                    row=row_number,
                    field="type",
                    message="type 必须是 ingredient 或 seasoning",
                ),
            )
        sort_order = _integer(row.get("sort_order"))
        if sort_order is None or sort_order < 0:
            recipe_error_keys.add(key)
            _add_issue(
                errors,
                RecipeImportIssue(
                    severity="error",
                    sheet="食材",
                    row=row_number,
                    field="sort_order",
                    message="sort_order 必须是大于等于 0 的整数",
                ),
            )
        _check_detail_title("食材", row, row_number, recipes_by_key[key], warnings)

    step_count_by_key: dict[str, int] = {}
    step_numbers_by_key: dict[str, set[int]] = {}
    for row_number, row in step_rows:
        key = _text(row.get("recipe_key"))
        if not key:
            _add_issue(
                errors,
                RecipeImportIssue(
                    severity="error",
                    sheet="步骤",
                    row=row_number,
                    field="recipe_key",
                    message="recipe_key 不能为空",
                ),
            )
            continue
        if key not in recipes_by_key:
            _add_issue(
                errors,
                RecipeImportIssue(
                    severity="error",
                    sheet="步骤",
                    row=row_number,
                    field="recipe_key",
                    message=f"找不到对应的菜谱：{key}",
                ),
            )
            continue

        step_count_by_key[key] = step_count_by_key.get(key, 0) + 1
        description = _text(row.get("description"))
        if not description:
            recipe_error_keys.add(key)
            _add_issue(
                errors,
                RecipeImportIssue(
                    severity="error",
                    sheet="步骤",
                    row=row_number,
                    field="description",
                    message="步骤描述不能为空",
                ),
            )

        step_number = _integer(row.get("step_number"))
        if step_number is None or step_number < 1:
            recipe_error_keys.add(key)
            _add_issue(
                errors,
                RecipeImportIssue(
                    severity="error",
                    sheet="步骤",
                    row=row_number,
                    field="step_number",
                    message="step_number 必须是大于等于 1 的整数",
                ),
            )
        elif step_number in step_numbers_by_key.setdefault(key, set()):
            recipe_error_keys.add(key)
            _add_issue(
                warnings,
                RecipeImportIssue(
                    severity="warning",
                    sheet="步骤",
                    row=row_number,
                    field="step_number",
                    message=f"同一道菜存在重复的步骤编号：{key} / {step_number}",
                ),
            )
        else:
            step_numbers_by_key[key].add(step_number)
        _check_detail_title("步骤", row, row_number, recipes_by_key[key], warnings)

    for recipe in recipes:
        key = recipe["recipe_key"]
        recipe["ingredient_count"] = ingredient_count_by_key.get(key, 0)
        recipe["step_count"] = step_count_by_key.get(key, 0)
        if recipe["ingredient_count"] == 0:
            recipe_error_keys.add(key)
            _add_issue(
                errors,
                RecipeImportIssue(
                    severity="error",
                    sheet="食材",
                    field="recipe_key",
                    message=f"菜谱没有食材明细：{key}",
                ),
            )
        if recipe["step_count"] == 0:
            recipe_error_keys.add(key)
            _add_issue(
                errors,
                RecipeImportIssue(
                    severity="error",
                    sheet="步骤",
                    field="recipe_key",
                    message=f"菜谱没有制作步骤：{key}",
                ),
            )

    if ingredient_unspecified_amount:
        _add_issue(
            warnings,
            RecipeImportIssue(
                severity="warning",
                sheet="食材",
                field="amount",
                message=f"有 {ingredient_unspecified_amount} 条食材用量为空或标记为“未标注”",
            ),
        )
    if ingredient_blank_unit:
        _add_issue(
            warnings,
            RecipeImportIssue(
                severity="warning",
                sheet="食材",
                field="unit",
                message=f"有 {ingredient_blank_unit} 条食材没有填写单位",
            ),
        )

    blank_cooking_time = sum(recipe["cooking_time"] is None for recipe in recipes)
    blank_image_url = sum(not recipe["image_url"] for recipe in recipes)
    if blank_cooking_time:
        _add_issue(
            warnings,
            RecipeImportIssue(
                severity="warning",
                sheet="菜谱",
                field="cooking_time",
                message=f"有 {blank_cooking_time} 道菜没有填写烹饪时间",
            ),
        )
    if blank_image_url:
        _add_issue(
            warnings,
            RecipeImportIssue(
                severity="warning",
                sheet="菜谱",
                field="image_url",
                message=f"有 {blank_image_url} 道菜没有图片地址",
            ),
        )
    if blank_tips:
        _add_issue(
            warnings,
            RecipeImportIssue(
                severity="warning",
                sheet="菜谱",
                field="tips/data_note",
                message=f"有 {blank_tips} 道菜没有填写 tips",
            ),
        )

    importable_count = sum(recipe["status"] == "可导入" for recipe in recipes)
    draft_count = sum(recipe["status"] == "草稿" for recipe in recipes)
    if recipes and importable_count == 0:
        _add_issue(
            warnings,
            RecipeImportIssue(
                severity="warning",
                sheet="菜谱",
                field="status",
                message="当前没有 status=可导入 的菜谱；本次仅可预览，不会有菜谱进入正式导入队列",
            ),
        )

    preview_items = [
        RecipeImportPreviewItem(**recipe)
        for recipe in recipes[:MAX_PREVIEW_RECIPES]
    ]
    return RecipeImportPreview(
        filename=filename,
        file_size_bytes=len(file_bytes),
        sheets=sheets,
        recipes_total=len(recipes),
        recipes_importable=importable_count,
        recipes_draft=draft_count,
        ingredient_rows=len(ingredient_rows),
        step_rows=len(step_rows),
        recipes=preview_items,
        errors=errors,
        warnings=warnings,
        can_import=bool(recipes and importable_count and not errors and not recipe_error_keys),
        truncated=len(recipes) > MAX_PREVIEW_RECIPES,
    )


def build_recipe_import_records(
    filename: str,
    file_bytes: bytes,
    include_drafts: bool = False,
) -> tuple[RecipeImportPreview, list[RecipeImportRecord]]:
    preview = preview_recipe_workbook(filename, file_bytes)
    if preview.errors:
        first_issue = preview.errors[0]
        raise ValueError(f"{first_issue.sheet}：{first_issue.message}")

    allowed_statuses = {"可导入"}
    if include_drafts:
        allowed_statuses.add("草稿")

    try:
        workbook = load_workbook(
            filename=BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
        recipe_rows, _ = _read_rows(workbook["菜谱"])
        ingredient_rows, _ = _read_rows(workbook["食材"])
        step_rows, _ = _read_rows(workbook["步骤"])
    except Exception as exc:
        raise ValueError("无法读取 Excel 明细，请确认文件没有损坏") from exc

    selected_keys = {
        _text(row.get("recipe_key"))
        for _, row in recipe_rows
        if _text(row.get("status")) in allowed_statuses
    }
    if not selected_keys:
        raise ValueError("没有符合导入条件的菜谱")

    ingredients_by_key: dict[str, list[RecipeIngredientBase]] = {}
    for _, row in ingredient_rows:
        key = _text(row.get("recipe_key"))
        if key not in selected_keys:
            continue
        ingredients_by_key.setdefault(key, []).append(
            RecipeIngredientBase(
                name=_text(row.get("name")) or "",
                amount=_text(row.get("amount")) or "",
                unit=_text(row.get("unit")) or "",
                type=_text(row.get("type")) or "ingredient",
                sort_order=_integer(row.get("sort_order")) or 0,
            )
        )

    steps_by_key: dict[str, list[RecipeStepBase]] = {}
    for _, row in step_rows:
        key = _text(row.get("recipe_key"))
        if key not in selected_keys:
            continue
        steps_by_key.setdefault(key, []).append(
            RecipeStepBase(
                step_number=_integer(row.get("step_number")) or 1,
                description=_text(row.get("description")) or "",
                duration=_text(row.get("duration")),
            )
        )

    records: list[RecipeImportRecord] = []
    for _, row in recipe_rows:
        key = _text(row.get("recipe_key"))
        if key not in selected_keys:
            continue
        tips_text = _text(row.get("tips")) or ""
        data = RecipeCreateRequest(
            title=_text(row.get("title")) or "",
            description=_text(row.get("description")) or "",
            category=_text(row.get("category")) or "其他",
            image_url=_text(row.get("image_url")),
            default_servings=_integer(row.get("default_servings")) or 2,
            cooking_time=_integer(row.get("cooking_time")),
            difficulty=_text(row.get("difficulty")) or "简单",
            tips=[line.strip() for line in tips_text.splitlines() if line.strip()],
            source_type=_text(row.get("source_type")) or "manual",
            source_url=_text(row.get("source_url")),
            ingredients=ingredients_by_key.get(key, []),
            steps=steps_by_key.get(key, []),
        )
        records.append(RecipeImportRecord(recipe_key=key, data=data))

    return preview, records


def import_recipe_records(
    db: Session,
    family_id: int,
    creator_id: int,
    records: list[RecipeImportRecord],
    filename: str,
) -> RecipeImportResult:
    backup = RecipeImportBackup(
        family_id=family_id,
        creator_id=creator_id,
        filename=filename,
        snapshot_json=_snapshot_existing_recipes(db, family_id, records),
        created_recipe_ids=[],
        imported_count=len(records),
    )
    db.add(backup)
    db.flush()

    result_items: list[RecipeImportResultItem] = []
    created_count = 0
    updated_count = 0
    created_recipe_ids: list[int] = []

    for record in records:
        statement = select(Recipe).where(
            Recipe.family_id == family_id,
            Recipe.recipe_key == record.recipe_key,
        )
        matches = list(db.execute(statement).scalars().all())
        if len(matches) > 1:
            raise ValueError(f"数据库中存在重复的 recipe_key：{record.recipe_key}")

        recipe = matches[0] if matches else None
        if recipe is None and record.data.image_url:
            legacy_statement = select(Recipe).where(
                Recipe.family_id == family_id,
                Recipe.recipe_key.is_(None),
                Recipe.image_url == record.data.image_url,
            )
            legacy_matches = list(db.execute(legacy_statement).scalars().all())
            if len(legacy_matches) > 1:
                raise ValueError(
                    f"数据库中存在多个相同图片地址的旧菜谱，无法安全匹配：{record.recipe_key}"
                )
            if legacy_matches:
                recipe = legacy_matches[0]
                recipe.recipe_key = record.recipe_key
        if recipe is None:
            recipe = Recipe(
                family_id=family_id,
                creator_id=creator_id,
                recipe_key=record.recipe_key,
            )
            db.add(recipe)
            action = "created"
            created_count += 1
        else:
            action = "updated"

        _apply_recipe_fields(recipe, record.data)
        _replace_ingredients(recipe, record.data)
        _replace_steps(recipe, record.data)
        if action == "created":
            db.flush()
            created_recipe_ids.append(recipe.id)
        result_items.append(
            RecipeImportResultItem(
                recipe_key=record.recipe_key,
                title=record.data.title,
                action=action,
                ingredient_count=len(record.data.ingredients),
                step_count=len(record.data.steps),
            )
        )

    backup.created_recipe_ids = created_recipe_ids
    db.commit()
    updated_count = len(result_items) - created_count
    return RecipeImportResult(
        filename=filename,
        imported_count=len(result_items),
        created_count=created_count,
        updated_count=updated_count,
        items=result_items,
        backup_id=backup.id,
        undo_available=True,
    )


def undo_recipe_import(
    db: Session,
    family_id: int,
    backup_id: int,
    creator_id: int,
) -> RecipeImportUndoResult:
    backup = db.execute(
        select(RecipeImportBackup).where(
            RecipeImportBackup.id == backup_id,
            RecipeImportBackup.family_id == family_id,
        )
    ).scalar_one_or_none()
    if backup is None:
        raise ValueError("找不到这次导入记录")
    if backup.creator_id != creator_id:
        raise ValueError("只有发起导入的人可以撤销这次导入")
    if backup.is_undone:
        raise ValueError("这次导入已经撤销过了")

    latest_backup_id = db.execute(
        select(RecipeImportBackup.id)
        .where(
            RecipeImportBackup.family_id == family_id,
            RecipeImportBackup.is_undone.is_(False),
        )
        .order_by(RecipeImportBackup.id.desc())
        .limit(1)
    ).scalar_one_or_none()
    if latest_backup_id != backup.id:
        raise ValueError("为避免覆盖更新，当前只能撤销最近一次导入")

    created_recipe_ids = {int(recipe_id) for recipe_id in backup.created_recipe_ids}
    references = _find_recipe_references(db, family_id, created_recipe_ids)
    if references:
        titles = {
            recipe_id: title
            for recipe_id, title in db.execute(
                select(Recipe.id, Recipe.title).where(Recipe.id.in_(references))
            ).all()
        }
        details = []
        for recipe_id, reference_types in sorted(references.items()):
            title = titles.get(recipe_id, f"菜谱 #{recipe_id}")
            details.append(f"「{title}」（{'、'.join(sorted(reference_types))}）")
        raise ValueError(
            "本次导入新增的菜谱已经被使用，不能撤销："
            + "；".join(details)
            + "。请先移除相关记录后再撤销。"
        )

    restored_count = 0
    for snapshot in backup.snapshot_json:
        recipe_id = int(snapshot["id"])
        recipe = _load_recipe_for_backup(db, family_id, recipe_id)
        if recipe is None:
            continue

        _restore_recipe_from_snapshot(recipe, snapshot)
        restored_count += 1

    removed_count = 0
    for recipe_id in created_recipe_ids:
        recipe = db.execute(
            select(Recipe).where(
                Recipe.id == int(recipe_id),
                Recipe.family_id == family_id,
            )
        ).scalar_one_or_none()
        if recipe is not None:
            db.delete(recipe)
            removed_count += 1

    if backup.snapshot_json:
        db.flush()
        _refresh_today_shopping_list_if_needed(
            db,
            family_id,
            {int(snapshot["id"]) for snapshot in backup.snapshot_json},
        )

    backup.is_undone = True
    backup.undone_at = datetime.now()
    db.commit()
    return RecipeImportUndoResult(
        filename=backup.filename,
        restored_count=restored_count,
        removed_count=removed_count,
    )


def _find_recipe_references(
    db: Session,
    family_id: int,
    recipe_ids: set[int],
) -> dict[int, set[str]]:
    """Find live and historical records that would be deleted with a recipe."""
    if not recipe_ids:
        return {}

    from app.models.daily_menu import DailyMenu, DailyMenuItem
    from app.models.daily_menu_feedback import DailyMenuFeedback
    from app.models.daily_menu_version import DailyMenuVersion
    from app.models.dish_order import DishOrder
    from app.models.recipe_activity import RecipeFavorite, RecipeViewHistory
    from app.models.weekly_menu import WeeklyMenuItem

    references: dict[int, set[str]] = {}

    def record(rows, label: str) -> None:
        for recipe_id in rows:
            references.setdefault(int(recipe_id), set()).add(label)

    record(
        db.execute(
            select(DishOrder.recipe_id).where(
                DishOrder.family_id == family_id,
                DishOrder.recipe_id.in_(recipe_ids),
            )
        ).scalars(),
        "点菜记录",
    )
    record(
        db.execute(
            select(DailyMenuItem.recipe_id)
            .join(DailyMenu, DailyMenu.id == DailyMenuItem.daily_menu_id)
            .where(
                DailyMenu.family_id == family_id,
                DailyMenuItem.recipe_id.in_(recipe_ids),
            )
        ).scalars(),
        "今日菜单",
    )
    record(
        db.execute(
            select(WeeklyMenuItem.recipe_id).where(
                WeeklyMenuItem.family_id == family_id,
                WeeklyMenuItem.recipe_id.in_(recipe_ids),
            )
        ).scalars(),
        "周菜单",
    )
    record(
        db.execute(
            select(DailyMenuFeedback.recipe_id).where(
                DailyMenuFeedback.family_id == family_id,
                DailyMenuFeedback.recipe_id.in_(recipe_ids),
            )
        ).scalars(),
        "成员反馈",
    )
    record(
        db.execute(
            select(RecipeFavorite.recipe_id).where(
                RecipeFavorite.recipe_id.in_(recipe_ids),
            )
        ).scalars(),
        "收藏记录",
    )
    record(
        db.execute(
            select(RecipeViewHistory.recipe_id).where(
                RecipeViewHistory.recipe_id.in_(recipe_ids),
            )
        ).scalars(),
        "浏览记录",
    )

    versions = db.execute(select(DailyMenuVersion).where(DailyMenuVersion.family_id == family_id)).scalars()
    for version in versions:
        for recipe_id in version.recipe_ids or []:
            try:
                normalized_recipe_id = int(recipe_id)
            except (TypeError, ValueError):
                continue
            if normalized_recipe_id in recipe_ids:
                references.setdefault(normalized_recipe_id, set()).add("今日菜单历史版本")

    return references


def _refresh_today_shopping_list_if_needed(
    db: Session,
    family_id: int,
    restored_recipe_ids: set[int],
) -> None:
    """Refresh today's generated list when an imported recipe was restored."""
    from app.models.daily_menu import DailyMenu, DailyMenuItem
    from app.models.family import Family
    from app.models.recipe import Recipe
    from app.services.shopping_list_service import rebuild_today_shopping_list

    today_menu = db.execute(
        select(DailyMenu)
        .options(
            selectinload(DailyMenu.items)
            .selectinload(DailyMenuItem.recipe)
            .selectinload(Recipe.ingredients)
        )
        .where(
            DailyMenu.family_id == family_id,
            DailyMenu.menu_date == datetime.now().date(),
            DailyMenu.status == "confirmed",
        )
    ).unique().scalar_one_or_none()
    if today_menu is None or not any(
        item.recipe_id in restored_recipe_ids for item in today_menu.items
    ):
        return

    family = db.get(Family, family_id)
    if family is not None:
        rebuild_today_shopping_list(db, family, today_menu)


def _snapshot_existing_recipes(
    db: Session,
    family_id: int,
    records: list[RecipeImportRecord],
) -> list[dict]:
    keys = {record.recipe_key for record in records}
    image_urls = {record.data.image_url for record in records if record.data.image_url}
    statement = (
        select(Recipe)
        .options(selectinload(Recipe.ingredients), selectinload(Recipe.steps))
        .where(Recipe.family_id == family_id)
    )
    recipes = db.execute(statement).unique().scalars().all()
    return [
        _serialize_recipe_snapshot(recipe)
        for recipe in recipes
        if recipe.recipe_key in keys
        or (recipe.recipe_key is None and recipe.image_url in image_urls)
    ]


def _serialize_recipe_snapshot(recipe: Recipe) -> dict:
    return {
        "id": recipe.id,
        "family_id": recipe.family_id,
        "creator_id": recipe.creator_id,
        "recipe_key": recipe.recipe_key,
        "title": recipe.title,
        "description": recipe.description,
        "category": recipe.category,
        "image_url": recipe.image_url,
        "default_servings": recipe.default_servings,
        "cooking_time": recipe.cooking_time,
        "difficulty": recipe.difficulty,
        "tips": list(recipe.tips or []),
        "source_type": recipe.source_type,
        "source_url": recipe.source_url,
        "ingredients": [
            {
                "name": item.name,
                "amount": item.amount,
                "unit": item.unit,
                "type": item.type,
                "sort_order": item.sort_order,
            }
            for item in recipe.ingredients
        ],
        "steps": [
            {
                "step_number": step.step_number,
                "description": step.description,
                "duration": step.duration,
            }
            for step in recipe.steps
        ],
    }


def _load_recipe_for_backup(db: Session, family_id: int, recipe_id: int) -> Recipe | None:
    statement = (
        select(Recipe)
        .options(selectinload(Recipe.ingredients), selectinload(Recipe.steps))
        .where(Recipe.id == recipe_id, Recipe.family_id == family_id)
    )
    return db.execute(statement).unique().scalar_one_or_none()


def _restore_recipe_from_snapshot(recipe: Recipe, snapshot: dict) -> None:
    data = RecipeCreateRequest(
        title=snapshot["title"],
        description=snapshot["description"],
        category=snapshot["category"],
        image_url=snapshot["image_url"],
        default_servings=snapshot["default_servings"],
        cooking_time=snapshot["cooking_time"],
        difficulty=snapshot["difficulty"],
        tips=snapshot["tips"],
        source_type=snapshot["source_type"],
        source_url=snapshot["source_url"],
        ingredients=snapshot["ingredients"],
        steps=snapshot["steps"],
    )
    _apply_recipe_fields(recipe, data)
    recipe.recipe_key = snapshot["recipe_key"]
    _replace_ingredients(recipe, data)
    _replace_steps(recipe, data)


def _read_rows(worksheet: Any) -> tuple[list[tuple[int, dict[str, Any]]], set[str]]:
    values = worksheet.iter_rows(values_only=True)
    try:
        header_values = next(values)
    except StopIteration:
        return [], set()

    headers = [_text(value) or "" for value in header_values]
    while headers and not headers[-1]:
        headers.pop()
    header_set = {header for header in headers if header}
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, row_values in enumerate(values, start=2):
        cells = list(row_values[: len(headers)])
        if not any(_text(value) for value in cells):
            continue
        cells.extend([None] * (len(headers) - len(cells)))
        rows.append((row_number, dict(zip(headers, cells))))
    return rows, header_set


def _check_headers(
    sheet: str,
    actual: set[str],
    required: set[str],
    errors: list[RecipeImportIssue],
) -> None:
    for header in sorted(required - actual):
        _add_issue(
            errors,
            RecipeImportIssue(
                severity="error",
                sheet=sheet,
                row=1,
                field=header,
                message=f"缺少必需列：{header}",
            ),
        )


def _validate_recipe_row(
    recipe: dict[str, Any],
    row_number: int,
    errors: list[RecipeImportIssue],
    warnings: list[RecipeImportIssue],
    recipe_error_keys: set[str],
) -> None:
    key = recipe["recipe_key"]
    title = recipe["title"]
    if len(title) < 2:
        recipe_error_keys.add(key)
        _add_issue(
            errors,
            RecipeImportIssue(
                severity="error",
                sheet="菜谱",
                row=row_number,
                field="title",
                message="title 至少需要 2 个字符",
            ),
        )
    if recipe["category"] not in RECIPE_CATEGORIES:
        recipe_error_keys.add(key)
        _add_issue(
            errors,
            RecipeImportIssue(
                severity="error",
                sheet="菜谱",
                row=row_number,
                field="category",
                message="category 不是软件支持的菜谱分类",
            ),
        )
    servings = recipe["default_servings"]
    if servings is None or not 1 <= servings <= 20:
        recipe_error_keys.add(key)
        _add_issue(
            errors,
            RecipeImportIssue(
                severity="error",
                sheet="菜谱",
                row=row_number,
                field="default_servings",
                message="default_servings 必须是 1 到 20 的整数",
            ),
        )
    cooking_time = recipe["cooking_time"]
    if recipe["cooking_time_raw"] and cooking_time is None:
        recipe_error_keys.add(key)
        _add_issue(
            errors,
            RecipeImportIssue(
                severity="error",
                sheet="菜谱",
                row=row_number,
                field="cooking_time",
                message="cooking_time 必须是整数分钟",
            ),
        )
    elif cooking_time is not None and not 1 <= cooking_time <= 1440:
        recipe_error_keys.add(key)
        _add_issue(
            errors,
            RecipeImportIssue(
                severity="error",
                sheet="菜谱",
                row=row_number,
                field="cooking_time",
                message="cooking_time 必须是 1 到 1440 分钟的整数",
            ),
        )
    if recipe["difficulty"] not in RECIPE_DIFFICULTIES:
        recipe_error_keys.add(key)
        _add_issue(
            errors,
            RecipeImportIssue(
                severity="error",
                sheet="菜谱",
                row=row_number,
                field="difficulty",
                message="difficulty 不是软件支持的难度",
            ),
        )
    if recipe["source_type"] not in RECIPE_SOURCE_TYPES:
        recipe_error_keys.add(key)
        _add_issue(
            errors,
            RecipeImportIssue(
                severity="error",
                sheet="菜谱",
                row=row_number,
                field="source_type",
                message="source_type 不是软件支持的来源类型",
            ),
        )
    if recipe["status"] not in RECIPE_STATUSES:
        recipe_error_keys.add(key)
        _add_issue(
            errors,
            RecipeImportIssue(
                severity="error",
                sheet="菜谱",
                row=row_number,
                field="status",
                message="status 必须是草稿或可导入",
            ),
        )
    if not recipe["image_url"]:
        _add_issue(
            warnings,
            RecipeImportIssue(
                severity="warning",
                sheet="菜谱",
                row=row_number,
                field="image_url",
                message="没有图片地址，导入后将使用无图占位",
            ),
        )


def _check_detail_title(
    sheet: str,
    row: dict[str, Any],
    row_number: int,
    recipe: dict[str, Any],
    warnings: list[RecipeImportIssue],
) -> None:
    detail_title = _text(row.get("recipe_title"))
    if detail_title and detail_title != recipe["title"]:
        _add_issue(
            warnings,
            RecipeImportIssue(
                severity="warning",
                sheet=sheet,
                row=row_number,
                field="recipe_title",
                message=f"recipe_title 与菜谱主表不一致，将以 recipe_key 对应的主表标题为准",
            ),
        )


def _text(value: Any) -> str | None:
    if value is None:
        return None
    return str(value).strip()


def _integer(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    try:
        text_value = str(value).strip()
        if text_value.endswith(".0"):
            text_value = text_value[:-2]
        return int(text_value)
    except (TypeError, ValueError):
        return None


def _add_issue(target: list[RecipeImportIssue], issue: RecipeImportIssue) -> None:
    if len(target) < MAX_ISSUES:
        target.append(issue)
